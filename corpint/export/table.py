import json
import unicodecsv as csv
from openpyxl import Workbook

from corpint.core import project, config
from corpint.model import Entity, Link, Mapping, Address, Document

MAPPING = {
    'entities': Entity,
    'links': Link,
    'addresses': Address,
    'documents': Document,
    'mappings': Mapping
}


def export_to_csv(relation, filename):
    """Export a single relation to a csv file."""
    model = MAPPING.get(relation)
    if model is None:
        project.log.info('No relation of type %s available.' % relation)
        return
    iter_rows = model.find()
    header = get_header(model)
    with open(filename, 'w+') as fh:
        w = csv.DictWriter(fh, header)
        w.writeheader()
        project.log.info('Writing CSV for relation %s...' % relation)
        for row in iter_rows:
            entity = {}
            for field in header:
                value = getattr(row, field)
                entity[field] = value
            w.writerow(entity)


def export_to_xlsx(filename):
    """Export all database entries to one combined xlsx file."""
    wb = Workbook()
    # Delete default sheet
    del wb['Sheet']
    for relation, model in MAPPING.items():
        make_sheet(wb, relation, model)
    wb.save(filename)


def make_sheet(wb, relation, model):
    """Create a sheet for each relation."""
    sheet = wb.create_sheet(relation)
    header = get_header(model)
    iter_rows = model.find()
    header_length = len(header)
    project.log.info('Adding sheet for relation %s...' % relation)
    data_fields = set()
    for idx, row in enumerate(iter_rows):
        row_number = idx + 1
        for jdx in range(header_length):
            col_number = jdx + 1
            value = header[jdx]
            if row_number > 1:
                value = getattr(row, header[jdx])
                if header[jdx] == 'data':
                    data = flatten_data(value)
                    for field, value in data.items():
                        data_fields.add(field)
                    continue
            sheet.cell(row=row_number, column=col_number, value=value)
    data_fields = list(data_fields)
    for idx, field in enumerate(data_fields):
        col_number = header_length + idx
        sheet.cell(row=1, column=col_number, value=field)
    for idx, row in enumerate(iter_rows):
        row_number = idx + 2
        if hasattr(row, 'data'):
            data = getattr(row, 'data')
            data = flatten_data(data)
            for field, value in data.items():
                field_idx = data_fields.index(field) + header_length
                if isinstance(value, list):
                    value = '; '.join(value)
                sheet.cell(row=row_number, column=field_idx, value=value)


def flatten_data(data):
    if 'data' in data.keys():
        data = data.get('data')
    return data


def get_header(model):
    return model.__table__.columns.keys()
